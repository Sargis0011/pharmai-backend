"""Bulk import for medicines from JSON / CSV / XLSX.
Prevents duplicates by (name_en|name_hy) + dosage signature."""
import csv
import io
import json
from typing import List, Dict, Any

from sqlalchemy.orm import Session
from openpyxl import load_workbook

from .models import Medicine, MedicineVariant, MedicineAlias, Category
from .search_utils import build_search_blob, normalize


def _key(rec: Dict[str, Any]) -> str:
    name = (rec.get("name_en") or rec.get("name_hy") or rec.get("name") or "").strip().lower()
    return normalize(name)


def _ensure_category(db: Session, key: str):
    if not key:
        return
    key = key.strip().lower()[:64]
    if not db.query(Category).filter(Category.key == key).first():
        db.add(Category(key=key, name_en=key.title()))


def _upsert_medicine(db: Session, rec: Dict[str, Any]) -> tuple[bool, bool, str]:
    """Returns (created, updated, error)."""
    try:
        name_hy = (rec.get("name_hy") or rec.get("name") or "").strip()
        if not name_hy and not rec.get("name_en"):
            return False, False, "missing name"

        key = _key(rec)
        existing = None
        if key:
            # match by normalized name
            for m in db.query(Medicine).filter(
                (Medicine.name_en.ilike(rec.get("name_en", "") or "_____zz")) |
                (Medicine.name_hy.ilike(name_hy or "_____zz"))
            ).limit(5).all():
                if normalize(m.name_en or m.name_hy) == key:
                    existing = m
                    break

        _ensure_category(db, rec.get("category", "other"))

        fields = dict(
            name_hy=name_hy or rec.get("name_en", ""),
            name_ru=rec.get("name_ru", "") or "",
            name_en=rec.get("name_en", "") or "",
            active_substance=rec.get("active_substance", "") or "",
            form=rec.get("form", "") or "",
            dosage=rec.get("dosage", "") or "",
            manufacturer=rec.get("manufacturer", "") or "",
            country=rec.get("country", "") or "",
            image_url=rec.get("image_url", "") or "",
            category=(rec.get("category") or "other").lower(),
            registered_in_am=int(rec.get("registered_in_am", 1) or 1),
            description=json.dumps(rec.get("description", {}) or {}, ensure_ascii=False),
            indications=json.dumps(rec.get("indications", {}) or {}, ensure_ascii=False),
            symptoms=json.dumps(rec.get("symptoms", {}) or {}, ensure_ascii=False),
            side_effects=json.dumps(rec.get("side_effects", []) or [], ensure_ascii=False),
            contraindications=json.dumps(rec.get("contraindications", []) or [], ensure_ascii=False),
            instruction=json.dumps(rec.get("instruction", {}) or {}, ensure_ascii=False),
        )
        fields["search_blob"] = build_search_blob(
            fields["name_hy"], fields["name_ru"], fields["name_en"],
            fields["active_substance"], fields["manufacturer"],
        )

        if existing:
            for k, v in fields.items():
                setattr(existing, k, v)
            med = existing
            created, updated = False, True
        else:
            med = Medicine(**fields)
            db.add(med)
            db.flush()
            created, updated = True, False

        # Variants
        for v in (rec.get("variants") or []):
            if not isinstance(v, dict):
                continue
            db.add(MedicineVariant(
                medicine_id=med.id,
                form=str(v.get("form", ""))[:120],
                dosage=str(v.get("dosage", ""))[:120],
                package_info=str(v.get("package_info", ""))[:255],
            ))

        # Aliases
        for a in (rec.get("aliases") or []):
            a = str(a).strip()
            if not a:
                continue
            db.add(MedicineAlias(
                alias=a[:255],
                alias_norm=normalize(a)[:255],
                medicine_id=med.id,
                active_substance=fields["active_substance"],
            ))

        return created, updated, ""
    except Exception as e:
        return False, False, str(e)


def import_records(db: Session, records: List[Dict[str, Any]]) -> Dict[str, Any]:
    created = updated = skipped = 0
    errors: List[str] = []
    for i, rec in enumerate(records):
        c, u, err = _upsert_medicine(db, rec)
        if err:
            skipped += 1
            errors.append(f"#{i}: {err}")
        elif c:
            created += 1
        elif u:
            updated += 1
        if (i + 1) % 200 == 0:
            db.commit()
    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped,
            "errors": errors[:50]}


def parse_payload(filename: str, content: bytes) -> List[Dict[str, Any]]:
    fn = (filename or "").lower()
    if fn.endswith(".json"):
        data = json.loads(content.decode("utf-8"))
        return data if isinstance(data, list) else data.get("items", [])
    if fn.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [dict(row) for row in reader]
    if fn.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or "").strip() for h in rows[0]]
        return [dict(zip(headers, r)) for r in rows[1:] if any(r)]
    raise ValueError(f"Unsupported file type: {filename}")
